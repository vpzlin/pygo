"""
Operate Excel File by openpyxl

Support Info:
    filename extension only supports ".xlsx"
Packages Required:
    pip install openpyxl     # offical document:  https://openpyxl.readthedocs.io/en/stable/
    pip install pillow       # to support images

"""


from openpyxl import Workbook, load_workbook
import src.common.File as File
import logging
import pandas as pd


class Excel:
    def __init__(self, file_path: str = None):
        # set logging
        logging.basicConfig(format='%(asctime)s | %(levelname)s | %(filename)s line:%(lineno)d | %(message)s',
                            level=logging.DEBUG)

        # variables
        self.__file_path: str = ""
        self.__workbook: Workbook = None
        self.__worksheet: Workbook.sheet = None

        # open excel file
        self.open_file(file_path)

    def __del__(self):
        pass

    """""""""""""""""""""""""""""""""""""""""""""""""""
    functions about excel file
    """""""""""""""""""""""""""""""""""""""""""""""""""
    def open_file(self, file_path: str):
        file_path = file_path.strip()
        # create new
        if file_path is None or len(file_path.strip()) == 0:
            self.__workbook = Workbook()
            self.__file_path = ""
            logging.debug("New Excel.")
            return True

        # file doesn't exist, create new file.
        if File.exists(file_path) is False:
            self.__workbook = Workbook()
            self.__file_path = file_path
            logging.info("New Excel. Path = [{0}]".format(file_path))
            return True

        if File.is_file(file_path) is False:
            logging.error("The file of the path isn't a file! Path = [{0}].".format(file_path))
            return False

        # open existed excel file
        try:
            self.__workbook = load_workbook(file_path)
            self.__file_path = file_path
            logging.info("Opened excel file. Path = [{0}].".format(file_path))
            return True
        except Exception as ex:
            logging.error("Failed to open excel file! Path = [{0}]. More info = [{1}]."
                          .format(file_path, str(ex)))
            return False

    def save_file(self):
        if len(self.__file_path) == 0:
            logging.error("This is a new created excel file! Please use function [save_file_as] ")
            return False

        try:
            self.__workbook.save(self.__file_path)
            logging.info("Saved excel file. Path = [{0}].".format(self.__file_path))
        except Exception as ex:
            logging.error("Failed to save excel file! More info = [{0}]".format(str(ex)))

    def save_file_as(self, file_path: str):
        file_path = file_path.strip()
        if len(file_path) == 0:
            logging.error("Fle path is illegal! Path = [{0}].".format(file_path))
            return False
        if File.exists(file_path):
            logging.error("Fle is already exist! Path = [{0}].".format(file_path))
            return False

        # save as a new file
        try:
            self.__workbook.save(file_path)
            if len(self.__file_path) == 0:
                self.__file_path = file_path
            logging.info("Saved excel as a new file. Path = [{0}].".format(self.__file_path))
            return True
        except Exception as ex:
            logging.error("Failed to save excel as a new file! Path = [{0}]. More info = [{1}]"
                          .format(file_path, str(ex)))
            return False

    """""""""""""""""""""""""""""""""""""""""""""""""""
    functions about excel sheet
    """""""""""""""""""""""""""""""""""""""""""""""""""
    def get_sheet_names(self):
        return self.__workbook.sheetnames

    def check_sheet_exist(self, sheet_name: str):
        if sheet_name.strip() in self.__workbook.sheetnames:
            return True
        return False

    def get_current_sheet_name(self):
        return self.__worksheet.title

    def open_sheet(self, sheet_name: str):
        sheet_name = sheet_name.strip()
        if sheet_name not in self.__workbook.sheetnames:
            logging.error("Failed to open sheet [{0}], it doesn't exist!".format(sheet_name))
            return False

        self.__worksheet = self.__workbook[sheet_name]
        logging.info("Opened sheet [{0}].".format(sheet_name))
        return True

    def add_sheet(self, sheet_name: str = None, index: int = None):
        if sheet_name is None:
            logging.error("Failed to create sheet! The sheet name [{0}] can't be none.".format(index))
            return False

        sheet_name = sheet_name.strip()
        if sheet_name in self.__workbook.sheetnames:
            logging.error("Failed to create sheet [{0}]! It's already existed.".format(sheet_name))
            return False

        try:
            self.__workbook.add_sheet(sheet_name, index)
            return True
        except Exception as ex:
            logging.error("Failed to create sheet [{0}]! More info = [{1}]."
                          .format(sheet_name, str(ex)))
            return False

    def drop_sheet(self, sheet_name: str):
        sheet_name = sheet_name.strip()
        if len(self.__workbook.sheetnames) == 1:
            logging.error("Failed to drop sheet [{0}]! There's only one sheet left in excel.".format(sheet_name))
            return False
        try:
            self.__workbook.remove(self.__workbook[sheet_name])
            logging.info("Dropped sheet [{0}].".format(sheet_name))
            return True
        except Exception as ex:
            logging.error("Failed to drop sheet [{0}]! More info = [{1}].".format(sheet_name, str(ex)))
            return False

    def copy_sheet(self, source_name: str, target_name: str):
        source_name = source_name.strip()
        target_name = target_name.strip()

        # check value
        if source_name not in self.__workbook.sheetnames:
            logging.error("Failed to rename sheet! Source sheet name [{0}] doesn't exist.".format(source_name))
            return False
        if target_name in self.__workbook.sheetnames:
            logging.error("Failed to rename sheet! Target sheet name [{0}] is already existed.".format(target_name))
            return False
        if source_name == target_name:
            logging.error("Source sheet name and target sheet name can't be set the same value [{0}]."
                          .format(source_name))
            return False

        # copy sheet
        try:
            ws = self.__workbook.copy_worksheet(self.__workbook[source_name])
            ws.title = target_name
            logging.info("Copied sheet from [{0}] to [{1}].".format(source_name, target_name))
            return True
        except Exception as ex:
            logging.error("Failed to copy sheet from [{0}] to [{1}]. More info = [{0}]."
                          .format(source_name, target_name, str(ex)))
            return False

    def rename_sheet(self, source_name: str, target_name: str):
        source_name = source_name.strip()
        target_name = target_name.strip()

        # check value
        if source_name not in self.__workbook.sheetnames:
            logging.error("Failed to rename sheet! Source sheet name [{0}] doesn't exist.".format(source_name))
            return False
        if target_name in self.__workbook.sheetnames:
            logging.error("Failed to rename sheet! Target sheet name [{0}] existes.".format(target_name))
            return False
        if source_name == target_name:
            logging.error("Source sheet name and target sheet name can't be set the same value [{0}]."
                          .format(source_name))
            return False

        # rename sheet
        try:
            if self.__worksheet is not None and self.__worksheet.title == source_name:
                self.__worksheet.title = target_name
            else:
                ws = self.__workbook[source_name]
                ws.title = target_name
            logging.info("Renamed sheet name from [{0}] to [{1}]."
                         .format(source_name, target_name))
            return True
        except Exception as ex:
            logging.error("Failed to rename sheet name from [{0}] to [{1}]! More info = [{2}]."
                          .format(source_name, target_name, str(ex)))
            return False

    """""""""""""""""""""""""""""""""""""""""""""""""""
    functions about data
    """""""""""""""""""""""""""""""""""""""""""""""""""
    def set_value(self, sheet_name: str, position: str, value: object):
        sheet_name = sheet_name.strip()
        position = position.strip().upper()
        if self.check_sheet_exist(sheet_name) is False:
            logging.error("Failed to set value on sheet [{0}]! It doesn't exist.".format(sheet_name))
            return False

        try:
            ws = self.__workbook[sheet_name]
            ws[position] = str(value)
            return True
        except Exception as ex:
            logging.error("Failed to set value on sheet [{0}]! More info = [{1}]."
                          .format(sheet_name, str(ex)))

    def get_value(self, sheet_name: str, position: str):
        sheet_name = sheet_name.strip()
        position = position.strip()
        if self.check_sheet_exist(sheet_name) is False:
            logging.error("Failed to set value on sheet [{0}]! It doesn't exist.".format(sheet_name))
            return ""
        try:
            ws = self.__workbook[sheet_name]
            if ws[position].value is None:
                return ""
            else:
                return str(ws[position].value)
        except Exception as ex:
            logging.error("Failed to set value on sheet [{0}]! More info = [{1}]."
                          .format(sheet_name, str(ex)))


def test():
    excel = Excel("C:\\GreenPark\\UserFiles\\Desktop\\Project\\Python\\pycoderx\\temp\\111.xlsx")
    #excel.open_sheet("Sheet")
    excel.set_value("Sheet", "A1", "It's A1.")
    print(excel.get_value("Sheet", "A1"))
    excel.save_file()


test()
